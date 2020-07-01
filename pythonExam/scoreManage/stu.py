import xlrd
import openpyxl
def selectStu(file1,file2,files3):
    # 打开文件
    f1 = xlrd.open_workbook(file1)
    f2 = xlrd.open_workbook(file2)
    # 输出第一个Sheet表格的名字
    print(f1.sheet_names()[0])
    print(f2.sheet_names()[0])
    # 获取第一个Sheet表格的名字
    sheetName1 = f1.sheet_names()[0]
    sheetName2 = f2.sheet_names()[0]
    # 根据名字获取第一个sheet表格的内容
    content1 = f1.sheet_by_name(sheetName1)
    content2 = f2.sheet_by_name(sheetName2)
    # 获取表格内容的所有行和列
    all_cols1 = content1.ncols  # 列
    all_rows1 = content1.nrows  # 行
    all_cols2 = content2.ncols  # 列
    all_rows2 = content2.nrows  # 行
    print(all_rows1)
    # 创建工作表
    wb = openpyxl.Workbook()
    # 定位工作表
    ws = wb.active
    index1  = 0
    index2 = 0
    for i in range(all_rows1):
        #在学生表中获取学生列表中的学号
        stuId1 = content1.cell(i,0).value
        stuName = content1.cell(i,1).value
        print(stuId1)
        #在所有学生中获取所有学生学号
        for j in range(all_rows2):
            stuId2 = content2.cell(j,2).value
            print(stuId2)
            #如果两个学生学号相等，将学号与成绩输入到新表中
            if stuId1 == stuId2:
                #获取学生成绩，并保留两位小数
                score1 = float(content2.cell(j,5).value)
                score1_1 = round(score1,2)
                score2 = float(content2.cell(j, 6).value)
                score2_1 = round(score2, 2)
                score3 = float(content2.cell(j, 7).value)
                score3_1 = round(score3, 2)
                score4 = float(content2.cell(j, 8).value)
                score4_1 = round(score4, 2)
                #将学号写入新表中
                pos = "A" + str(i+1)
                ws[pos] = stuId1
                #将学生姓名写入新表中
                pos1 = "B" + str(i + 1)
                ws[pos1] = stuName
                pos2 = "C" + str(i + 1)
                pos3 = "D" + str(i + 1)
                pos4 = "E" + str(i + 1)
                pos5 = "F" + str(i + 1)
                ws[pos2] = score1_1
                ws[pos3] = score2_1
                ws[pos4] = score3_1
                ws[pos5] = score4_1
                #跳出本次循环
                continue
    wb.save(files3)
files1 = input("请输入班级Excel绝对地址：\n")
files2 = input("请输入所有成绩Excel表的绝对地址：\n")
files3 = input("请输入保存的绝对地址:\n")
selectStu(files1,files2,files3)
#selectStu("F:\lht\Desktop\\znwl.xlsx","F:\lht\Desktop\\2019-2020.xls")
